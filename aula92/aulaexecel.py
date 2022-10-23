import openpyxl
from random import uniform

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Fabio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Marco {linha} {round(uniform(10, 100), 2)}'

planilha.save('trabalho.xlsx')